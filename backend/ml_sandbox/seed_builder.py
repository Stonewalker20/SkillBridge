"""Build synthetic SkillBridge documents from local seed datasets."""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from bson import ObjectId


def _now() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def normalize_skill_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _token_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", normalize_skill_text(value))
    return slug.strip("-") or "seed"


def _clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _dedupe(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _clean_text(value)
        key = text.lower()
        if not text or key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


@dataclass(frozen=True)
class SkillSeed:
    name: str
    category: str
    aliases: tuple[str, ...]


@dataclass(frozen=True)
class EvidenceSeed:
    user_key: str
    user_email: str
    type: str
    title: str
    source: str
    text_excerpt: str
    tags: tuple[str, ...]


@dataclass(frozen=True)
class ResumeSeed:
    raw_text: str
    category: str
    source: str


@dataclass(frozen=True)
class JobSeed:
    title: str
    company: str
    location: str
    source: str
    description_excerpt: str
    required_skills: tuple[str, ...]


@dataclass(frozen=True)
class SeedSources:
    skills: tuple[SkillSeed, ...]
    evidence: tuple[EvidenceSeed, ...]
    resumes: tuple[ResumeSeed, ...]
    jobs: tuple[JobSeed, ...]
    manifest: dict[str, Any]


def _match_skill_ids(text: str, skill_docs: list[dict[str, Any]]) -> list[str]:
    lowered = normalize_skill_text(text)
    matched: list[str] = []
    for skill in skill_docs:
        names = [skill["name"], *(skill.get("aliases") or [])]
        for name in names:
            token = normalize_skill_text(name)
            if not token:
                continue
            pattern = rf"(?<![a-z0-9]){re.escape(token)}(?![a-z0-9])"
            if re.search(pattern, lowered):
                matched.append(str(skill["_id"]))
                break
    return _dedupe(matched)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _open_zipped_csv_rows(archive_path: Path, member_name: str) -> Iterable[dict[str, str]]:
    with zipfile.ZipFile(archive_path) as archive:
        with archive.open(member_name, "r") as raw:
            wrapper = io.TextIOWrapper(raw, encoding="utf-8", errors="ignore", newline="")
            reader = csv.DictReader(wrapper)
            for row in reader:
                yield {str(key or ""): str(value or "") for key, value in row.items()}


def _load_seed_skills(seed_dir: Path) -> list[SkillSeed]:
    rows = _read_json(seed_dir / "skills.json")
    return [
        SkillSeed(
            name=_clean_text(row.get("name") or ""),
            category=_clean_text(row.get("category") or "General"),
            aliases=tuple(_dedupe(str(alias or "") for alias in (row.get("aliases") or []))),
        )
        for row in rows
        if _clean_text(row.get("name") or "")
    ]


def _load_seed_evidence(seed_dir: Path) -> list[EvidenceSeed]:
    rows = _read_json(seed_dir / "evidence.json")
    out: list[EvidenceSeed] = []
    for index, row in enumerate(rows, start=1):
        email = _clean_text(row.get("user_email") or f"seed-evidence-{index}@example.com")
        out.append(
            EvidenceSeed(
                user_key=email.lower(),
                user_email=email,
                type=_clean_text(row.get("type") or "project"),
                title=_clean_text(row.get("title") or f"Seed Evidence {index}"),
                source=_clean_text(row.get("source") or "seed"),
                text_excerpt=_clean_text(row.get("text_excerpt") or ""),
                tags=tuple(_dedupe(str(value or "") for value in (row.get("tags") or []))),
            )
        )
    return out


def _load_seed_jobs(seed_dir: Path) -> list[JobSeed]:
    rows = _read_json(seed_dir / "jobs.json")
    out: list[JobSeed] = []
    for index, row in enumerate(rows, start=1):
        out.append(
            JobSeed(
                title=_clean_text(row.get("title") or f"Seed Job {index}"),
                company=_clean_text(row.get("company") or "SkillBridge Seed Co"),
                location=_clean_text(row.get("location") or "Remote"),
                source=_clean_text(row.get("source") or "seed"),
                description_excerpt=_clean_text(row.get("description_excerpt") or ""),
                required_skills=tuple(_dedupe(str(value or "") for value in (row.get("required_skills") or []))),
            )
        )
    return out


def _load_resume_zip_rows(seed_dir: Path, max_rows: int) -> list[ResumeSeed]:
    archive_path = seed_dir / "resume-dataset.zip"
    if not archive_path.exists() or max_rows <= 0:
        return []
    out: list[ResumeSeed] = []
    for row in _open_zipped_csv_rows(archive_path, "Resume/Resume.csv"):
        raw_text = _clean_text(row.get("Resume_str") or "")
        if len(raw_text) < 200:
            continue
        out.append(
            ResumeSeed(
                raw_text=raw_text,
                category=_clean_text(row.get("Category") or "General"),
                source="resume-dataset.zip",
            )
        )
        if len(out) >= max_rows:
            break
    return out


def _canonical_required_skills(raw_skills: Iterable[str], skill_names: set[str]) -> list[str]:
    out: list[str] = []
    for value in raw_skills:
        skill = _clean_text(value)
        normalized = normalize_skill_text(skill)
        if not skill:
            continue
        if normalized in skill_names:
            out.append(skill)
    return _dedupe(out)


def _extract_skill_names_from_text(text: str, skill_names: set[str]) -> list[str]:
    lowered = normalize_skill_text(text)
    out: list[str] = []
    for skill_name in skill_names:
        pattern = rf"(?<![a-z0-9]){re.escape(skill_name)}(?![a-z0-9])"
        if re.search(pattern, lowered):
            out.append(skill_name)
    return _dedupe(out)


def _load_external_postings(seed_dir: Path, skill_names: set[str], max_rows: int) -> list[JobSeed]:
    archive_path = seed_dir / "linkedin-job-postings.zip"
    if not archive_path.exists() or max_rows <= 0:
        return []
    out: list[JobSeed] = []
    for row in _open_zipped_csv_rows(archive_path, "postings.csv"):
        title = _clean_text(row.get("title") or "")
        description = _clean_text(row.get("description") or "")
        if not title or len(description) < 160:
            continue
        required = _canonical_required_skills(
            [part.strip() for part in re.split(r"[,;/|]", row.get("skills_desc") or "")],
            skill_names,
        )
        if not required:
            required = _extract_skill_names_from_text(description, skill_names)
        if not required:
            continue
        out.append(
            JobSeed(
                title=title,
                company=_clean_text(row.get("company_name") or "LinkedIn Seed Company"),
                location=_clean_text(row.get("location") or "Remote"),
                source="linkedin-job-postings.zip",
                description_excerpt=description[:900],
                required_skills=tuple(required),
            )
        )
        if len(out) >= max_rows:
            break
    return out


def _load_large_linkedin_jobs(seed_dir: Path, skill_names: set[str], max_rows: int) -> list[JobSeed]:
    archive_path = seed_dir / "1-3m-linkedin-jobs-and-skills-2024.zip"
    if not archive_path.exists() or max_rows <= 0:
        return []

    matched_skill_map: dict[str, list[str]] = {}
    for row in _open_zipped_csv_rows(archive_path, "job_skills.csv"):
        job_link = _clean_text(row.get("job_link") or "")
        raw_skills = [part.strip() for part in re.split(r",", row.get("job_skills") or "")]
        required = _canonical_required_skills(raw_skills, skill_names)
        if job_link and required:
            matched_skill_map[job_link] = required
        if len(matched_skill_map) >= max_rows * 3:
            break

    out: list[JobSeed] = []
    for row in _open_zipped_csv_rows(archive_path, "linkedin_job_postings.csv"):
        job_link = _clean_text(row.get("job_link") or "")
        required = matched_skill_map.get(job_link)
        if not required:
            continue
        title = _clean_text(row.get("job_title") or "")
        if not title:
            continue
        description = (
            f"Role focus: {title}. "
            f"Primary skills: {', '.join(required)}. "
            f"This posting was synthesized from LinkedIn job and skill metadata for ML sandbox seeding."
        )
        out.append(
            JobSeed(
                title=title,
                company=_clean_text(row.get("company") or "LinkedIn Seed Company"),
                location=_clean_text(row.get("job_location") or "Remote"),
                source="1-3m-linkedin-jobs-and-skills-2024.zip",
                description_excerpt=description,
                required_skills=tuple(required),
            )
        )
        if len(out) >= max_rows:
            break
    return out


def _load_nyc_jobs(seed_dir: Path, skill_names: set[str], max_rows: int) -> list[JobSeed]:
    archive_path = seed_dir / "nyc-job-postings-with-esco-occupations-and-skills.zip"
    if not archive_path.exists() or max_rows <= 0:
        return []

    from openpyxl import load_workbook
    import tempfile

    out: list[JobSeed] = []
    with zipfile.ZipFile(archive_path) as archive:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            jobs_name = "NYC Jobs (snapshot 2025-07-13) (anonymized).xlsx"
            skills_name = "skills - 15000 sample.xlsx"
            archive.extract(jobs_name, temp_path)
            archive.extract(skills_name, temp_path)
            jobs_wb = load_workbook(temp_path / jobs_name, read_only=True)
            skills_wb = load_workbook(temp_path / skills_name, read_only=True)

            jobs_ws = jobs_wb[jobs_wb.sheetnames[0]]
            rows = jobs_ws.iter_rows(values_only=True)
            header = [str(value or "") for value in next(rows)]
            desired_jobs: list[dict[str, Any]] = []
            desired_ids: set[str] = set()
            for row in rows:
                payload = {header[index]: row[index] for index in range(min(len(header), len(row)))}
                description = _clean_text(payload.get("description") or "")
                if len(description) < 160:
                    continue
                job_id = str(payload.get("id") or "").strip()
                if not job_id:
                    continue
                desired_jobs.append(payload)
                desired_ids.add(job_id)
                if len(desired_jobs) >= max_rows * 3:
                    break

            skills_ws = skills_wb[skills_wb.sheetnames[0]]
            skill_rows = skills_ws.iter_rows(values_only=True)
            skill_header = [str(value or "") for value in next(skill_rows)]
            skills_by_job: dict[str, list[str]] = {}
            for row in skill_rows:
                payload = {skill_header[index]: row[index] for index in range(min(len(skill_header), len(row)))}
                job_id = str(payload.get("jobposting_id") or "").strip()
                if job_id not in desired_ids:
                    continue
                label = _clean_text(payload.get("skill_esco_label") or payload.get("skill_esco_alias") or "")
                if normalize_skill_text(label) in skill_names:
                    skills_by_job.setdefault(job_id, []).append(label)

            for payload in desired_jobs:
                job_id = str(payload.get("id") or "").strip()
                required = _dedupe(skills_by_job.get(job_id, []))
                if not required:
                    continue
                out.append(
                    JobSeed(
                        title=_clean_text(payload.get("Title") or payload.get("Civil Service Title") or "NYC Seed Job"),
                        company=_clean_text(payload.get("company_name") or "NYC Government"),
                        location="New York, NY",
                        source="nyc-job-postings-with-esco-occupations-and-skills.zip",
                        description_excerpt=_clean_text(payload.get("description") or "")[:900],
                        required_skills=tuple(required),
                    )
                )
                if len(out) >= max_rows:
                    break
    return out


def load_seed_sources(
    seed_dir: str | Path,
    *,
    max_resume_rows: int = 12,
    max_external_postings: int = 12,
    max_large_linkedin_jobs: int = 12,
    max_nyc_jobs: int = 12,
) -> SeedSources:
    root = Path(seed_dir)
    skills = _load_seed_skills(root)
    skill_names = {normalize_skill_text(skill.name) for skill in skills}
    evidence = _load_seed_evidence(root)
    resumes = _load_resume_zip_rows(root, max_rows=max_resume_rows)
    jobs = [
        *_load_seed_jobs(root),
        *_load_external_postings(root, skill_names, max_rows=max_external_postings),
        *_load_large_linkedin_jobs(root, skill_names, max_rows=max_large_linkedin_jobs),
        *_load_nyc_jobs(root, skill_names, max_rows=max_nyc_jobs),
    ]
    manifest = {
        "seed_dir": str(root),
        "loaded_counts": {
            "skills": len(skills),
            "evidence": len(evidence),
            "resume_rows": len(resumes),
            "job_templates": len(jobs),
        },
        "source_breakdown": {
            "manual_jobs": sum(1 for job in jobs if job.source in {"hand-collected", "linkedin"}),
            "resume_zip": len(resumes),
            "linkedin_postings_zip": sum(1 for job in jobs if job.source == "linkedin-job-postings.zip"),
            "large_linkedin_zip": sum(1 for job in jobs if job.source == "1-3m-linkedin-jobs-and-skills-2024.zip"),
            "nyc_zip": sum(1 for job in jobs if job.source == "nyc-job-postings-with-esco-occupations-and-skills.zip"),
        },
    }
    return SeedSources(
        skills=tuple(skills),
        evidence=tuple(evidence),
        resumes=tuple(resumes),
        jobs=tuple(jobs),
        manifest=manifest,
    )


def build_seed_documents(
    seed_sources: SeedSources,
    *,
    seed_namespace: str = "ml-sandbox-seed",
    max_jobs_per_user: int = 2,
) -> dict[str, Any]:
    now = _now()
    skill_docs: list[dict[str, Any]] = []
    skill_id_by_name: dict[str, ObjectId] = {}
    for index, skill in enumerate(seed_sources.skills):
        oid = ObjectId()
        skill_docs.append(
            {
                "_id": oid,
                "name": skill.name,
                "category": skill.category,
                "aliases": list(skill.aliases),
                "origin": "synthetic-seed",
                "hidden": False,
                "seed_namespace": seed_namespace,
                "created_at": now + timedelta(seconds=index),
                "updated_at": now + timedelta(seconds=index),
            }
        )
        skill_id_by_name[normalize_skill_text(skill.name)] = oid
        for alias in skill.aliases:
            skill_id_by_name.setdefault(normalize_skill_text(alias), oid)

    def skill_ids_for_terms(values: Iterable[str]) -> list[ObjectId]:
        out: list[ObjectId] = []
        seen: set[str] = set()
        for value in values:
            oid = skill_id_by_name.get(normalize_skill_text(value))
            if oid is None:
                continue
            key = str(oid)
            if key in seen:
                continue
            seen.add(key)
            out.append(oid)
        return out

    users: list[dict[str, Any]] = []
    user_id_by_key: dict[str, ObjectId] = {}
    user_email_by_key: dict[str, str] = {}
    user_counter = 0

    def ensure_user(key: str, label: str | None = None) -> ObjectId:
        nonlocal user_counter
        if key in user_id_by_key:
            return user_id_by_key[key]
        user_counter += 1
        oid = ObjectId()
        username = f"seed_user_{user_counter:03d}"
        email = f"{username}@example.com"
        user_id_by_key[key] = oid
        user_email_by_key[key] = email
        users.append(
            {
                "_id": oid,
                "email": email,
                "username": username,
                "role": "user",
                "is_active": True,
                "seed_label": label or username,
                "seed_namespace": seed_namespace,
                "created_at": now + timedelta(minutes=user_counter),
            }
        )
        return oid

    evidence_docs: list[dict[str, Any]] = []
    resume_snapshots: list[dict[str, Any]] = []
    confirmations: list[dict[str, Any]] = []
    profile_skill_ids: dict[str, list[ObjectId]] = {}
    evidence_by_user: dict[str, list[dict[str, Any]]] = {}

    for index, row in enumerate(seed_sources.evidence, start=1):
        user_oid = ensure_user(row.user_key, label=row.user_email)
        tag_skill_ids = skill_ids_for_terms(row.tags)
        inferred_skill_ids = skill_ids_for_terms(
            [
                *row.tags,
                *[
                    skill["name"]
                    for skill in skill_docs
                    if str(skill["_id"]) in _match_skill_ids(row.text_excerpt, skill_docs)
                ],
            ]
        )
        skill_ids = tag_skill_ids or inferred_skill_ids
        evidence_doc = {
            "_id": ObjectId(),
            "user_id": user_oid,
            "user_email": user_email_by_key[row.user_key],
            "type": row.type,
            "title": row.title,
            "source": row.source,
            "text_excerpt": row.text_excerpt,
            "skill_ids": skill_ids,
            "origin": "synthetic",
            "seed_namespace": seed_namespace,
            "created_at": now + timedelta(hours=index),
            "updated_at": now + timedelta(hours=index),
        }
        evidence_docs.append(evidence_doc)
        evidence_by_user.setdefault(str(user_oid), []).append(evidence_doc)
        profile_skill_ids.setdefault(str(user_oid), [])
        profile_skill_ids[str(user_oid)].extend(skill_ids)

        if row.type.lower() == "resume":
            snapshot_oid = ObjectId()
            raw_text = (
                f"{row.title}\nSUMMARY\n{row.text_excerpt}\nSKILLS\n"
                f"{', '.join(_dedupe(row.tags))}"
            )
            resume_snapshots.append(
                {
                    "_id": snapshot_oid,
                    "user_id": user_oid,
                    "source_type": "seed_resume_excerpt",
                    "raw_text": raw_text,
                    "metadata": {"source": row.source},
                    "image_ref": "/images/resume_icon.png",
                    "seed_namespace": seed_namespace,
                    "created_at": now + timedelta(hours=index, minutes=5),
                }
            )
            if skill_ids:
                confirmations.append(
                    {
                        "_id": ObjectId(),
                        "user_id": user_oid,
                        "resume_snapshot_id": snapshot_oid,
                        "confirmed": [
                            {"skill_id": skill_id, "skill_name": next(skill["name"] for skill in skill_docs if skill["_id"] == skill_id), "proficiency": 3, "manual_proficiency": 3}
                            for skill_id in skill_ids
                        ],
                        "rejected": [],
                        "edited": [],
                        "seed_namespace": seed_namespace,
                        "created_at": now + timedelta(hours=index, minutes=6),
                        "updated_at": now + timedelta(hours=index, minutes=6),
                    }
                )

    for index, row in enumerate(seed_sources.resumes, start=1):
        key = f"resume:{index}"
        user_oid = ensure_user(key, label=row.category)
        matched_ids = [ObjectId(value) for value in _match_skill_ids(row.raw_text, skill_docs)]
        snapshot_oid = ObjectId()
        resume_snapshots.append(
            {
                "_id": snapshot_oid,
                "user_id": user_oid,
                "source_type": "seed_resume_zip",
                "raw_text": row.raw_text[:4000],
                "metadata": {"source": row.source, "category": row.category},
                "image_ref": "/images/resume_icon.png",
                "seed_namespace": seed_namespace,
                "created_at": now + timedelta(days=index),
            }
        )
        if matched_ids:
            confirmations.append(
                {
                    "_id": ObjectId(),
                    "user_id": user_oid,
                    "resume_snapshot_id": snapshot_oid,
                    "confirmed": [
                        {"skill_id": skill_id, "skill_name": next(skill["name"] for skill in skill_docs if skill["_id"] == skill_id), "proficiency": 3, "manual_proficiency": 3}
                        for skill_id in matched_ids
                    ],
                    "rejected": [],
                    "edited": [],
                    "seed_namespace": seed_namespace,
                    "created_at": now + timedelta(days=index, minutes=1),
                    "updated_at": now + timedelta(days=index, minutes=1),
                }
            )
            evidence_doc = {
                "_id": ObjectId(),
                "user_id": user_oid,
                "user_email": user_email_by_key[key],
                "type": "resume",
                "title": f"{row.category.title()} Resume Snapshot",
                "source": row.source,
                "text_excerpt": row.raw_text[:600],
                "skill_ids": matched_ids,
                "origin": "synthetic",
                "seed_namespace": seed_namespace,
                "created_at": now + timedelta(days=index, minutes=2),
                "updated_at": now + timedelta(days=index, minutes=2),
            }
            evidence_docs.append(evidence_doc)
            evidence_by_user.setdefault(str(user_oid), []).append(evidence_doc)
            profile_skill_ids.setdefault(str(user_oid), [])
            profile_skill_ids[str(user_oid)].extend(matched_ids)

    for key, values in list(profile_skill_ids.items()):
        profile_skill_ids[key] = skill_ids_for_terms(
            [next(skill["name"] for skill in skill_docs if skill["_id"] == value) for value in values]
        )

    job_templates = list(seed_sources.jobs)
    job_ingests: list[dict[str, Any]] = []
    job_match_runs: list[dict[str, Any]] = []
    tailored_resumes: list[dict[str, Any]] = []

    for user_index, user in enumerate(users, start=1):
        user_oid = user["_id"]
        user_key = str(user_oid)
        user_skills = profile_skill_ids.get(user_key, [])
        user_skill_names = {
            next(skill["name"] for skill in skill_docs if skill["_id"] == skill_id)
            for skill_id in user_skills
        }
        scored_templates: list[tuple[int, int, JobSeed]] = []
        for template_index, template in enumerate(job_templates):
            required_ids = skill_ids_for_terms(template.required_skills)
            overlap = len({str(skill_id) for skill_id in required_ids} & {str(skill_id) for skill_id in user_skills})
            scored_templates.append((overlap, -template_index, template))
        scored_templates.sort(reverse=True)
        selected_templates = [template for _score, _index, template in scored_templates[: max(1, max_jobs_per_user)]]
        for template_index, template in enumerate(selected_templates, start=1):
            required_ids = skill_ids_for_terms(template.required_skills)
            extracted_skills = [
                {"skill_id": str(skill_id), "skill_name": next(skill["name"] for skill in skill_docs if skill["_id"] == skill_id), "matched_on": "seed", "count": 1}
                for skill_id in required_ids
            ]
            job_text = (
                f"{template.title} at {template.company}. "
                f"Location: {template.location}. "
                f"{template.description_excerpt} "
                f"Required skills: {', '.join(template.required_skills)}."
            ).strip()
            job_oid = ObjectId()
            job_doc = {
                "_id": job_oid,
                "user_id": user_oid,
                "title": template.title,
                "company": template.company,
                "location": template.location,
                "text": job_text,
                "extracted_skills": extracted_skills,
                "keywords": [normalize_skill_text(value) for value in template.required_skills][:12],
                "seed_namespace": seed_namespace,
                "created_at": now + timedelta(days=user_index, hours=template_index),
            }
            job_ingests.append(job_doc)

            matched_evidence = [
                evidence
                for evidence in evidence_by_user.get(user_key, [])
                if {str(skill_id) for skill_id in (evidence.get("skill_ids") or [])} & {str(skill_id) for skill_id in required_ids}
            ]
            if not matched_evidence:
                matched_evidence = evidence_by_user.get(user_key, [])[:2]
            retrieved_context = [
                {
                    "source_type": "evidence",
                    "source_id": str(evidence["_id"]),
                    "title": evidence["title"],
                    "snippet": evidence["text_excerpt"][:220],
                    "score": round(0.65 + (0.08 * index), 2),
                    "chunk_index": 0,
                }
                for index, evidence in enumerate(matched_evidence[:3], start=0)
            ]
            selected_skill_ids = required_ids[:8] or user_skills[:8]
            selected_item_ids = [evidence["_id"] for evidence in matched_evidence[:3]]
            matched_skill_names = _dedupe(
                [
                    *[
                        next(skill["name"] for skill in skill_docs if skill["_id"] == skill_id)
                        for skill_id in selected_skill_ids
                    ],
                    *list(user_skill_names),
                ]
            )[:10]
            bullet_lines = [
                f"- {_clean_text(evidence['text_excerpt'])[:140].rstrip('.')}."
                for evidence in matched_evidence[:3]
                if _clean_text(evidence.get("text_excerpt") or "")
            ]
            if not bullet_lines:
                bullet_lines = [f"- Demonstrated experience with {', '.join(matched_skill_names[:3])} in prior work."]
            tailored_resumes.append(
                {
                    "_id": ObjectId(),
                    "user_id": user_oid,
                    "job_id": job_oid,
                    "resume_snapshot_id": None,
                    "template_source": "seed_template",
                    "job_text": job_text,
                    "template": "seed_v1",
                    "selected_skill_ids": selected_skill_ids,
                    "selected_item_ids": selected_item_ids,
                    "retrieved_context": retrieved_context,
                    "sections": [
                        {"title": "Summary", "lines": [f"Targeted for {template.title} with emphasis on {', '.join(matched_skill_names[:4])}."]},
                        {"title": "Targeted Highlights", "lines": bullet_lines},
                    ],
                    "plain_text": "\n".join(["Summary", f"Targeted for {template.title}.", "Targeted Highlights", *bullet_lines]),
                    "seed_namespace": seed_namespace,
                    "created_at": now + timedelta(days=user_index, hours=template_index, minutes=10),
                    "updated_at": now + timedelta(days=user_index, hours=template_index, minutes=10),
                }
            )
            job_match_runs.append(
                {
                    "_id": ObjectId(),
                    "user_id": user_oid,
                    "job_id": job_oid,
                    "title": template.title,
                    "company": template.company,
                    "location": template.location,
                    "text_preview": job_text[:220],
                    "job_text_snapshot": job_text,
                    "analysis": {
                        "job_id": str(job_oid),
                        "matched_skills": matched_skill_names[:6],
                        "retrieved_context": retrieved_context,
                        "template_source": "seed_template",
                    },
                    "seed_namespace": seed_namespace,
                    "created_at": now + timedelta(days=user_index, hours=template_index, minutes=5),
                    "updated_at": now + timedelta(days=user_index, hours=template_index, minutes=5),
                }
            )

    manifest = {
        "generated_at": now.isoformat(),
        "seed_namespace": seed_namespace,
        "source_manifest": seed_sources.manifest,
        "counts": {
            "users": len(users),
            "skills": len(skill_docs),
            "evidence": len(evidence_docs),
            "resume_snapshots": len(resume_snapshots),
            "confirmations": len(confirmations),
            "job_ingests": len(job_ingests),
            "job_match_runs": len(job_match_runs),
            "tailored_resumes": len(tailored_resumes),
        },
    }
    return {
        "users": users,
        "skills": skill_docs,
        "evidence": evidence_docs,
        "resume_snapshots": resume_snapshots,
        "resume_skill_confirmations": confirmations,
        "job_ingests": job_ingests,
        "job_match_runs": job_match_runs,
        "tailored_resumes": tailored_resumes,
        "manifest": manifest,
    }
